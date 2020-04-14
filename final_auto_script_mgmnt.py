#!/usr/bin/python

import MySQLdb
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
import re
from lxml import etree
from datetime import date, timedelta


db = MySQLdb.connect("172.17.3.33","imis","opossum","plp" )
sql_workbook = xlsxwriter.Workbook('sql_report.xlsx')
sql_worksheet = sql_workbook.add_worksheet()
row = 0
col = 0
cursor = db.cursor()
yesterday = date.today() - timedelta(1)
dt = yesterday.strftime('%Y-%m-%d')
dt = "2018-04-27"
sqlcommand = "select count(*),i_application_master.status,kubera.k_institutions.inst_name from i_application_master LEFT OUTER join i_netbanking_fetch_details on i_netbanking_fetch_details.master_id=i_application_master.id  join kubera.k_institutions on k_institutions.inst_id = i_netbanking_fetch_details.institution_id where (i_application_master.processing_type = 'BANK_ACCOUNT_DETAILS_FETCH' or i_application_master.processing_type = 'NETBANKING_FETCH') and (date(i_netbanking_fetch_details.created) = '%s' or date(i_netbanking_fetch_details.updated) = '%s') group by i_application_master.status,kubera.k_institutions.inst_name order by inst_name" % (dt,dt)
try:
    cursor.execute(sqlcommand) 
    result = cursor.fetchall() 
    for r in result:
        sql_worksheet.write(row, col, r[0])
        sql_worksheet.write(row, col+1, r[1])
        sql_worksheet.write(row, col+2, r[2])
        row+=1
except:
    print "Error : Unable to fetch data"
db.close()
sql_workbook.close()


wb = load_workbook('sql_report.xlsx')
ws = wb.active

wb2 = load_workbook('gopi_report.xlsx')
ws2 = wb2.active

workbook = xlsxwriter.Workbook('combined_report.xlsx')
worksheet = workbook.add_worksheet()

workbook2 = xlsxwriter.Workbook('final.xlsx')
worksheet2 = workbook2.add_worksheet()


first_column = ws['A']
second_column = ws['B']
third_column = ws['C']

f_row = len(ws2[1])



instList = []
statusList = []
errorList = []
insights_data = {}
scripts_data = {}
combined_data = {}
combined_data1 = {}
itrv_data = {}
final_data = {}

for x in xrange(len(third_column)): 
    instList.append(third_column[x].value)
for x in xrange(len(second_column)): 
    statusList.append(second_column[x].value)
for x in xrange(2,f_row+1) :
    value = ws2.cell(row=1,column=x).value
    errorList.append(value)


rowLt = len(second_column)
colLt = len(ws[1])
for i in xrange (1,rowLt+1) :
    insights_data[i] = {}
    for j in xrange (1,colLt+1) :
        value = ws.cell(row=i, column=j).value   
        insights_data[i][j] = value

        
rowLt2 = len(ws2['A'])
colLt2 = len(ws2[1])
for i in xrange (1,rowLt2+1) :
    scripts_data[i] = {}
    for j in xrange (1,colLt2+1) :
        value = ws2.cell(row=i, column=j).value   
        scripts_data[i][j] = value


def Unique(tempList) :
    finalList = []
    for num in tempList:
        if num not in finalList:
            finalList.append(num)
    return finalList
    
instList = Unique(instList)
instList.sort()
ins = len(instList)
statusList = Unique(statusList)
statusList.sort()
sta = len(statusList)
err = len(errorList)

r=1
c=0
for value in instList :
    worksheet.write(r,c,value)
    r+=1

r=1
c=0
worksheet2.write(0,0,"Inst Name")
for value in instList :
    worksheet2.write(r,c,value)
    r+=1

ro = 0
co = 1
for status in statusList :
    worksheet.write(ro,co,status)
    co+=1

sa = co
for error in errorList :
    worksheet.write(ro,co,error)
    co+=1
workbook.close();
    
wbook = load_workbook('combined_report.xlsx')
wsheet = wbook.active
    
#Database Report    
for i in xrange (1,rowLt+1) :
    temp = insights_data[i][3]
    for a in xrange (0,ins) :
        if temp == instList[a] :
            temp2 = insights_data[i][2]
            for b in xrange (0,sta) :
                if temp2 == statusList[b] :
                    value1 = wsheet.cell(row=a+2,column=b+2).value
                    if value1 == None :
                        value1 = 0
                    values = insights_data[i][1]
                    values = values + float(value1)
                    wsheet.cell(a+1,b+1).value = values
                    

#Gopi report
for i in xrange (1,rowLt2+1) :
    temp = scripts_data[i][1]
    for a in xrange (0,ins) :
        if temp == instList[a] :
            for b in xrange (1,err+1) :
                j = b + 1
                value = scripts_data[i][j]
                y = j + sa -2
                wsheet.cell(a+1,y).value = value
                  
                    

    

workbook2.close()
wb.save('sql_report.xlsx')
wb2.save('gopi_report.xlsx')
wbook.save('combined_report.xlsx')


wb3 = load_workbook('combined_report.xlsx')
ws3 = wb3.active

#Reading joined data (both gopi and sql)
rowLt3 = len(ws3['A'])
colLt3 = len(ws3[1])
for i in xrange (1,rowLt3+1) :
    combined_data[i] = {}
    for j in xrange (1,colLt3+1) :
        value = ws3.cell(row=i, column=j).value   
        combined_data[i][j] = value

#Insights total attempts
ws3.cell(1,colLt3+1).value = 'INSIGHTS TOTAL ATTEMPTS'
for i in xrange (2,rowLt3+1) :
    result = 0
    for b in xrange (2,sta+2) :
        value = combined_data[i][b]
        if value == None :
            value = 0
        result = result + value
    ws3.cell(i,colLt3+1).value = result

#Scripts total attempts
ws3.cell(1,colLt3+2).value = 'SCRIPTS TOTAL ATTEMPTS'
for i in xrange (2,rowLt3+1) :
    result = 0
    for b in xrange (1,3) :
        b = b + colLt3 - 2
        value = combined_data[i][b]
        if value == None :
            value = 0
        result = result + value
    ws3.cell(i,colLt3+2).value = result


#Reading joined data2 (both gopi and sql)
colLt4 = len(ws3[1])
for i in xrange (1,rowLt3+1) :
    combined_data1[i] = {}
    for j in xrange (1,colLt4+1) :
        value = ws3.cell(row=i, column=j).value   
        combined_data1[i][j] = value

#Final xslx sheet
fb = load_workbook('final.xlsx')
fs = fb.active

fbc = len(fs[1])

#Insights total attempts
fs.cell(1,fbc+1).value = 'INSIGHTS TOTAL ATTEMPTS'
for i in xrange (2,rowLt3+1) :
    result = 0
    for b in xrange (2,sta+2) :
        value = combined_data[i][b]
        if value == None :
            value = 0
        result = result + value
    fs.cell(i,fbc+1).value = result

fbc1 = len(fs[1])

#Scripts total attempts
fs.cell(1,fbc1+1).value = 'SCRIPTS TOTAL ATTEMPTS'
for i in xrange (2,rowLt3+1) :
    result = 0
    for b in xrange (1,3) :
        b = b + colLt3 - 2
        value = combined_data[i][b]
        if value == None :
            value = 0
        result = result + value
    fs.cell(i,fbc1+1).value = result

fbc2 = len(fs[1])

#Did not proceed
fs.cell(1,fbc2+1).value = 'SCRIPTS NOT PROCEED'
for i in xrange (2,rowLt3+1) :
    result = combined_data1[i][colLt4-1] - combined_data1[i][colLt4]
    fs.cell(i,fbc2+1).value = result

fbc3 = len(fs[1])

#Script Success
fs.cell(1,fbc3+1).value = 'SCRIPT SUCCESS'
sucIndex = errorList.index('TOTAL SUCCESS')
sucIndex = sucIndex + sta + 2
for i in xrange (2,rowLt3+1) :
    temp = combined_data1[i][sucIndex]
    fs.cell(i,fbc3+1).value = temp

fbc4 =  len(fs[1])

#Script failure
fs.cell(1,fbc4+1).value = 'SCRIPT FAILURE'
for i in xrange (2,rowLt3+1) :
    if combined_data1[i][colLt4] == None :
        combined_data1[i][colLt4] = 0
    if combined_data1[i][sucIndex] == None :
        combined_data1[i][sucIndex] = 0
    result = combined_data1[i][colLt4] - combined_data1[i][sucIndex]
    fs.cell(i,fbc4+1).value = result


fbc5 = len(fs[1])

#User Errors
fs.cell(1,fbc5+1).value = 'USER ERRORS'
userIndex = errorList.index('USERACT') + sta + 2
chpasIndex = errorList.index('CHPASSWORD') + sta + 2
noaccIndex = errorList.index('NOACC') + sta + 2
passIndex = errorList.index('PASSWORD') + sta + 2
for i in xrange (2,rowLt3+1) :
    result = combined_data1[i][userIndex] + combined_data1[i][chpasIndex] + combined_data1[i][noaccIndex] + combined_data1[i][passIndex]
    fs.cell(i,fbc5+1).value = result

fbc6 = len(fs[1])

#Site Errors
fs.cell(1,fbc6+1).value = 'SITE ERRORS'
siteIndex = errorList.index('SITE') + sta + 2
httpIndex = errorList.index('HTTP') + sta + 2
timeoutIndex = errorList.index('TIMEOUT') + sta + 2
pageIndex = errorList.index('PAGE') + sta + 2
noerrorIndex = errorList.index('NO ERROR CODE') + sta + 2
for i in xrange (2,rowLt3+1) :
    result = combined_data1[i][siteIndex] + combined_data1[i][httpIndex] + combined_data1[i][timeoutIndex] + combined_data1[i][pageIndex] + combined_data1[i][noerrorIndex]
    fs.cell(i,fbc6+1).value = result
fbc7 = len(fs[1])

#Geniune and Fixable site errors
fs.cell(1,fbc7+1).value = 'GENIUNE SITE ERRORS'
fs.cell(1,fbc7+2).value = 'FIXABLE SITE ERRORS'
for i in xrange (1,rowLt3+1) :
    final_data[i] = {}
    for j in xrange (1,fbc7+1) :
        value = fs.cell(row=i, column=j).value   
        final_data[i][j] = value
for i in xrange (2,rowLt3+1) :
    fs.cell(i,fbc7+2).value = 0
    fs.cell(i,fbc7+1).value = fs.cell(i,fbc7).value - fs.cell(i,fbc7+2).value
fbc8 = len(fs[1])

#Form 26AS
workbook = xlsxwriter.Workbook('itrv_report.xlsx')
worksheet = workbook.add_worksheet()
fname = "/media/finch/itrFailures/bheema/"+"summary"+"."+dt+".html"
f = open(fname,"r")
r = 0
if f.mode=='r' :
    content = f.read()
table = etree.HTML(content).find("body/table")
rows = iter(table)
headers = [col.text for col in next(rows)]
for row in rows:
    values = [col.text for col in row]
    for i in xrange (0,3) :
        worksheet.write(r, i, values[i])
    r+=1
workbook.close()
f.close()
wb4 = load_workbook('itrv_report.xlsx')
ws4 = wb4.active 
#Reading data
rEnd = len(ws4['A'])
cEnd = len(ws4[1])
for i in xrange (1,rEnd+1) :
    itrv_data[i] = {}
    for j in xrange (1,cEnd+1) :
        value = ws4.cell(row=i, column=j).value   
        itrv_data[i][j] = value

itrvList = {"InsightsYesBankComboFetch" : "Yes Bank" , "InsightsCorporationBankCombo" : "Corporation Bank" , "InsightsDenaBankComboFetch" : "Dena Bank" , "InsightsKotakbankCombo" : "Kotak Mahindra / Ing Vysya Bank" , "InsightsStateBankCombo" : "State Bank of India" , "InsightsUnionBankComboFetch" : "Union Bank of India" , "InsightsUCOCombo" : "UCO Bank" , "InsightsHDFCComboFetch" : "HDFC Bank" , "InsightsICICIComboFetch" : "ICICI Bank" , "InsightsIDBIComboFetch" : "IDBI" ,  "InsightsCitibankCombo" : "Citibank" ,  "InsightsAXISComboFetch" : "Axis Bank" }
#Success,total attempts and failure count
fs.cell(1,fbc8+1).value = 'FORM26AS/ITRV ATTEMPTS'
fs.cell(1,fbc8+2).value = 'FORM26AS/ITRV SUCCESS'
fs.cell(1,fbc8+3).value = 'FORM26AS/ITRV ERRORS'
for key in itrvList :
    for i in xrange (1,rEnd+1) :
        temp = itrv_data[i][1]
        if temp == key :
            total = itrv_data[i][2]
            total = int(total)
            success = itrv_data[i][3]
            if success == None :
                success = 0
            success = int(success)
            failure = total - success
            value = itrvList[key]
            for i in xrange (2,rowLt3+1) :
                temp2 = combined_data1[i][1]
                if temp2 == value :
                    fs.cell(i,fbc8+1).value = total
                    fs.cell(i,fbc8+2).value = success
                    fs.cell(i,fbc8+3).value = failure  
print "Executed.."

wb3.save('combined_report.xlsx')
wb4.save('itrv_report.xlsx')
fb.save('final.xlsx')
